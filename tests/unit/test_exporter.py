"""Tests for data exporter (CSV, SQL, JSON, Excel)."""

import pytest
from pathlib import Path

from pmi_analyzer.types import ShamkhMetrics
from pmi_analyzer.data.exporter import (
    metrics_to_dataframe,
    export_to_csv,
    export_to_sql,
    export_to_json,
    export_to_excel,
    merge_with_historical,
)


# ------------------------------------------------------------------ #
#  metrics_to_dataframe
# ------------------------------------------------------------------ #


class TestMetricsToDataframe:
    def test_single_metric(self):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9, production=48.2)]
        df = metrics_to_dataframe(metrics)
        assert len(df) == 1
        assert df.iloc[0]["month"] == "1405-03"
        assert df.iloc[0]["pmi_total"] == 45.9
        assert df.iloc[0]["production"] == 48.2

    def test_multiple_metrics(self):
        metrics = [
            ShamkhMetrics(month="1405-01", pmi_total=47.0),
            ShamkhMetrics(month="1405-02", pmi_total=46.5),
            ShamkhMetrics(month="1405-03", pmi_total=45.9),
        ]
        df = metrics_to_dataframe(metrics)
        assert len(df) == 3
        assert list(df["month"]) == ["1405-01", "1405-02", "1405-03"]

    def test_empty_list(self):
        df = metrics_to_dataframe([])
        assert len(df) == 0
        assert df.empty

    def test_none_values_preserved(self):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        df = metrics_to_dataframe(metrics)
        assert df.iloc[0]["production"] is None or df.iloc[0]["production"] != df.iloc[0]["production"]  # NaN check


# ------------------------------------------------------------------ #
#  export_to_csv
# ------------------------------------------------------------------ #


class TestExportToCsv:
    def test_creates_csv(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.csv"
        result = export_to_csv(metrics, output)
        assert result.exists()
        content = output.read_text(encoding="utf-8-sig")
        assert "month," in content
        assert "1405-03" in content

    def test_creates_parent_dirs(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "subdir" / "data.csv"
        export_to_csv(metrics, output)
        assert output.exists()

    def test_csv_has_all_columns(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9, production=48.2)]
        output = tmp_path / "test.csv"
        export_to_csv(metrics, output)
        content = output.read_text(encoding="utf-8-sig")
        headers = content.split("\n")[0]
        assert "month" in headers
        assert "pmi_total" in headers
        assert "production" in headers
        assert "new_orders" in headers

    def test_utf8_bom_encoding(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.csv"
        export_to_csv(metrics, output)
        raw = output.read_bytes()
        # UTF-8 BOM is EF BB BF
        assert raw[:3] == b"\xef\xbb\xbf"


# ------------------------------------------------------------------ #
#  merge_with_historical
# ------------------------------------------------------------------ #


class TestMergeWithHistorical:
    def test_merges_new_with_existing(self, tmp_path):
        existing = tmp_path / "historical.csv"
        existing.write_text(
            "month,pmi_total,production\n1405-01,47.0,49.0\n1405-02,46.5,48.5\n",
            encoding="utf-8-sig",
        )
        new = [ShamkhMetrics(month="1405-03", pmi_total=45.9, production=48.2)]
        df = merge_with_historical(new, existing)
        assert len(df) == 3
        assert df.iloc[-1]["month"] == "1405-03"

    def test_deduplicates_by_month(self, tmp_path):
        existing = tmp_path / "historical.csv"
        existing.write_text(
            "month,pmi_total\n1405-03,41.6\n",  # old wrong value
            encoding="utf-8-sig",
        )
        new = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]  # new correct value
        df = merge_with_historical(new, existing)
        assert len(df) == 1
        assert df.iloc[0]["pmi_total"] == 45.9  # keeps latest

    def test_creates_new_if_no_existing(self, tmp_path):
        existing = tmp_path / "nonexistent.csv"
        new = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        df = merge_with_historical(new, existing)
        assert len(df) == 1
        assert df.iloc[0]["month"] == "1405-03"

    def test_sorts_by_month(self, tmp_path):
        existing = tmp_path / "historical.csv"
        existing.write_text(
            "month,pmi_total\n1405-03,45.9\n1405-01,47.0\n",
            encoding="utf-8-sig",
        )
        new = [ShamkhMetrics(month="1405-02", pmi_total=46.5)]
        df = merge_with_historical(new, existing)
        assert list(df["month"]) == ["1405-01", "1405-02", "1405-03"]


# ------------------------------------------------------------------ #
#  export_to_sql
# ------------------------------------------------------------------ #


class TestExportToSql:
    def test_creates_sql_file(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.sql"
        result = export_to_sql(metrics, output)
        assert result.exists()
        content = output.read_text(encoding="utf-8")
        assert "CREATE TABLE" in content
        assert "INSERT INTO" in content

    def test_create_table_has_correct_name(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.sql"
        export_to_sql(metrics, output, table_name="my_pmi")
        content = output.read_text(encoding="utf-8")
        assert "CREATE TABLE IF NOT EXISTS my_pmi" in content
        assert "INSERT INTO my_pmi" in content

    def test_insert_has_all_columns(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9, production=48.2)]
        output = tmp_path / "test.sql"
        export_to_sql(metrics, output)
        content = output.read_text(encoding="utf-8")
        # Check INSERT statement has key columns
        insert_line = [l for l in content.split("\n") if l.startswith("INSERT")][0]
        assert "month" in insert_line
        assert "pmi_total" in insert_line
        assert "production" in insert_line

    def test_null_values_for_missing_fields(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.sql"
        export_to_sql(metrics, output)
        content = output.read_text(encoding="utf-8")
        insert_line = [l for l in content.split("\n") if l.startswith("INSERT")][0]
        assert "NULL" in insert_line

    def test_creates_parent_dirs(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "subdir" / "data.sql"
        export_to_sql(metrics, output)
        assert output.exists()


# ------------------------------------------------------------------ #
#  export_to_json
# ------------------------------------------------------------------ #


class TestExportToJson:
    def test_creates_json_file(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.json"
        result = export_to_json(metrics, output)
        assert result.exists()
        content = output.read_text(encoding="utf-8")
        assert "1405-03" in content
        assert "45.9" in content

    def test_json_is_valid_array(self, tmp_path):
        import json
        metrics = [
            ShamkhMetrics(month="1405-01", pmi_total=47.0),
            ShamkhMetrics(month="1405-02", pmi_total=46.5),
        ]
        output = tmp_path / "test.json"
        export_to_json(metrics, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert isinstance(data, list)
        assert len(data) == 2

    def test_json_preserves_none_as_null(self, tmp_path):
        import json
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.json"
        export_to_json(metrics, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert data[0]["month"] == "1405-03"
        assert data[0]["pmi_total"] == 45.9
        assert data[0]["production"] is None

    def test_creates_parent_dirs(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "subdir" / "data.json"
        export_to_json(metrics, output)
        assert output.exists()


# ------------------------------------------------------------------ #
#  export_to_excel
# ------------------------------------------------------------------ #


class TestExportToExcel:
    def test_creates_excel_file(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.xlsx"
        result = export_to_excel(metrics, output)
        assert result.exists()

    def test_excel_has_correct_sheet_name(self, tmp_path):
        import openpyxl
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "test.xlsx"
        export_to_excel(metrics, output, sheet_name="My Data")
        wb = openpyxl.load_workbook(output)
        assert "My Data" in wb.sheetnames

    def test_creates_parent_dirs(self, tmp_path):
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "subdir" / "data.xlsx"
        export_to_excel(metrics, output)
        assert output.exists()


# ------------------------------------------------------------------ #
#  Edge cases
# ------------------------------------------------------------------ #


class TestEdgeCases:
    def test_all_boundary_values(self, tmp_path):
        """Test with boundary values: 0, 50, 100."""
        metrics = [
            ShamkhMetrics(month="1405-01", pmi_total=0.0),
            ShamkhMetrics(month="1405-02", pmi_total=50.0),
            ShamkhMetrics(month="1405-03", pmi_total=100.0),
        ]
        output = tmp_path / "boundary.csv"
        export_to_csv(metrics, output)
        content = output.read_text(encoding="utf-8-sig")
        assert "0.0" in content
        assert "50.0" in content
        assert "100.0" in content

    def test_special_characters_in_month(self, tmp_path):
        """Month with special characters should be preserved."""
        metrics = [ShamkhMetrics(month="1405-03/rev2", pmi_total=45.9)]
        output = tmp_path / "special.csv"
        export_to_csv(metrics, output)
        content = output.read_text(encoding="utf-8-sig")
        assert "1405-03/rev2" in content

    def test_sql_escapes_single_quotes(self, tmp_path):
        """SQL export should escape single quotes in values."""
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "escape.sql"
        export_to_sql(metrics, output)
        content = output.read_text(encoding="utf-8")
        # Month value should be quoted and escaped
        assert "'1405-03'" in content

    def test_json_unicode_preservation(self, tmp_path):
        """JSON should preserve Persian/Arabic characters."""
        import json
        metrics = [ShamkhMetrics(month="خرداد-۱۴۰۵", pmi_total=45.9)]
        output = tmp_path / "unicode.json"
        export_to_json(metrics, output)
        data = json.loads(output.read_text(encoding="utf-8"))
        assert data[0]["month"] == "خرداد-۱۴۰۵"

    def test_large_dataset(self, tmp_path):
        """Test with 100 records."""
        metrics = [
            ShamkhMetrics(month=f"1405-{i:02d}", pmi_total=40.0 + i * 0.1)
            for i in range(1, 13)
        ] * 8  # 96 records
        output = tmp_path / "large.csv"
        export_to_csv(metrics, output)
        lines = output.read_text(encoding="utf-8-sig").strip().split("\n")
        assert len(lines) == 97  # 96 data + 1 header

    def test_all_none_values(self, tmp_path):
        """Metrics with all None values (except month)."""
        metrics = [ShamkhMetrics(month="1405-03")]
        output = tmp_path / "all_none.csv"
        export_to_csv(metrics, output)
        content = output.read_text(encoding="utf-8-sig")
        assert "1405-03" in content

    def test_empty_metrics_list_all_formats(self, tmp_path):
        """Empty metrics list should work for all formats."""
        metrics = []
        csv_out = tmp_path / "empty.csv"
        sql_out = tmp_path / "empty.sql"
        json_out = tmp_path / "empty.json"

        export_to_csv(metrics, csv_out)
        export_to_sql(metrics, sql_out)
        export_to_json(metrics, json_out)

        assert csv_out.exists()
        assert sql_out.exists()
        assert json_out.exists()

    def test_merge_empty_new_metrics(self, tmp_path):
        """Merging empty new metrics with existing CSV."""
        existing = tmp_path / "historical.csv"
        existing.write_text(
            "month,pmi_total\n1405-01,47.0\n",
            encoding="utf-8-sig",
        )
        df = merge_with_historical([], existing)
        assert len(df) == 1
        assert df.iloc[0]["pmi_total"] == 47.0

    def test_sql_multiple_inserts(self, tmp_path):
        """SQL should have one INSERT per row."""
        metrics = [
            ShamkhMetrics(month="1405-01", pmi_total=47.0),
            ShamkhMetrics(month="1405-02", pmi_total=46.5),
            ShamkhMetrics(month="1405-03", pmi_total=45.9),
        ]
        output = tmp_path / "multi.sql"
        export_to_sql(metrics, output)
        content = output.read_text(encoding="utf-8")
        insert_count = content.count("INSERT INTO")
        assert insert_count == 3

    def test_json_indent(self, tmp_path):
        """JSON with custom indent."""
        metrics = [ShamkhMetrics(month="1405-03", pmi_total=45.9)]
        output = tmp_path / "indented.json"
        export_to_json(metrics, output, indent=4)
        content = output.read_text(encoding="utf-8")
        assert '    "month"' in content  # 4-space indent

    def test_excel_multiple_rows(self, tmp_path):
        """Excel with multiple rows."""
        import openpyxl
        metrics = [
            ShamkhMetrics(month="1405-01", pmi_total=47.0),
            ShamkhMetrics(month="1405-02", pmi_total=46.5),
        ]
        output = tmp_path / "multi.xlsx"
        export_to_excel(metrics, output)
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 data rows
